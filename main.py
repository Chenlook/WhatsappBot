from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook

def SendMessage(messages):
    for message in messages:
        input_message = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true'][data-tab='10']")))
        input_message.click()
        sleep(0.2)

        lines = message.split('\n')
        for line in lines:
            input_message.send_keys(line)
            input_message.send_keys(Keys.SHIFT, Keys.ENTER)  # saut de ligne sans envoyer
        sleep(0.2)

        input_message.send_keys(Keys.ENTER)  # envoyer le message

def ouvrir_whatsapp():
    global driver, wait, chrome_options

    chrome_options = Options()
    chrome_options.add_argument("user-data-dir=C:\\Users\\chaca\\AppData\\Local\\Google\\Chrome\\SeleniumProfile")
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 30)

    driver.get("https://web.whatsapp.com")


def MainFunction(users, messages):
    for user in users:
        driver.get(f"https://wa.me/{user}")

        sleep(1)

        continue_button = wait.until(EC.element_to_be_clickable((By.ID, "action-button")))
        continue_button.click()

        sleep(1)

        use_whatsapp = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "utilisez WhatsApp Web")))
        use_whatsapp.click()

        sleep(1)

        SendMessage(messages)

        sleep(2)

def import_file():
    global path

    file_path = filedialog.askopenfilename(title="Select a file", filetypes=[("Text files", "*.xlsx"), ("All files", "*.*")])
    if file_path:
        # Process the selected file (you can replace this with your own logic)
        path = file_path

def recuperer_colonne():
    global column

    column = column_input.get()

def recuperer_message():
    global message

    message = message_input.get() 

def cleanNum(numero):
    newNum = "33"
    for i in range(1,len(numero)):
        if numero[i] != " ":
            newNum += numero[i]
    return newNum

def verifNum(numero):
    if numero[0] == "0":
        if len(numero) == 14:
            return True
    return False

    
def Programme():
    global liste_contact, message, driver, wait, chrome_options

    # charger les numéros ici
    wb = load_workbook(path)
    ws = wb.active
    liste_contact = []
    for cell in ws[column]:
        if verifNum(str(cell.value)):
            liste_contact.append(cleanNum(str(cell.value)))

    print("Contacts valides :", liste_contact)

    # récupérer le message depuis le Text widget
    message = message_input.get("1.0", "end-1c")

    chrome_options = Options()
    chrome_options.add_argument("user-data-dir=C:\\Users\\chaca\\AppData\\Local\\Google\\Chrome\\SeleniumProfile")
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 30)

    MainFunction(liste_contact, [message])

    sleep(2)

    driver.quit()


    

def init_window():
    global column_input, message_input

    maFenetre = Tk()
    maFenetre.geometry("600x440")
    maFenetre.title("WhatsApp bot")


    connect_button = Button(maFenetre, text="Se connecter à WhatsApp", command=ouvrir_whatsapp)
    connect_button.pack(pady=10)

    top_frame = Frame(maFenetre, bg="lightgrey", padx=10, pady=10)
    top_frame.pack(side="top", fill="x")
    import_text = Label(top_frame, text="Importer un fichier excel")
    import_button = Button(top_frame, text="Importer", command=import_file)
    column_text = Label(top_frame, text="Nom de la colonne des numéros?")
    column_input = Entry(top_frame, width=40)
    column_button = Button(top_frame, text="Valider", command=recuperer_colonne)
    import_text.grid(row=1, column=0, sticky="w",pady=20, padx=10)
    import_button.grid(row=1, column=1, sticky="w",pady=20, padx=20)
    column_text.grid(row=2, column=0, sticky="w",pady=20, padx=10)
    column_input.grid(row=2, column=1, sticky="w",pady=20, padx=20)
    column_button.grid(row=2, column=2, sticky="w",pady=20)

    message_input = Text(maFenetre, width=40, height=8)
    message_input.pack(pady=20)

    start_button = Button(maFenetre, text="Commencer", command=Programme)
    start_button.pack(pady=20)

    maFenetre.mainloop()

column = ""
path = ""
message = ""
liste_contact = []


init_window()


